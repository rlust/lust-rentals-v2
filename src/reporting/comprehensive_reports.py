"""Comprehensive Multi-Phase Reporting System for Lust Rentals."""

import os
import logging
from typing import Dict, Optional
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.data_processing.processor import FinancialDataProcessor
from src.utils.config import load_config
from src.categorization.category_utils import normalize_category, get_display_name

logger = logging.getLogger(__name__)


class ComprehensiveReportGenerator:
    """Generates Phase 1 Excel reports with LLC summary, expense matrix, and property details."""

    def __init__(self, data_dir: Path = None):
        """Initialize the report generator."""
        config = load_config()
        self.data_dir = data_dir or config.data_dir
        self.processor = FinancialDataProcessor(data_dir=self.data_dir)
        self.reports_dir = os.path.join(self.data_dir, "reports")
        os.makedirs(self.reports_dir, exist_ok=True)

    def generate_phase1_excel(self, year: int, output_file: str = None) -> str:
        """Generate comprehensive Excel report for Phase 1."""
        logger.info(f"Generating Phase 1 Excel report for {year}")
        
        result = self.processor.load_processed_data(year)
        income_df = result["income"].copy()
        expenses_df = result["expenses"].copy()
        
        if not expenses_df.empty and 'category' in expenses_df.columns:
            expenses_df['category_display'] = expenses_df['category'].apply(
                lambda x: get_display_name(normalize_category(x))
            )
        
        wb = Workbook()
        wb.remove(wb.active)
        
        all_properties = sorted(set(
            list(income_df['property_name'].unique() if not income_df.empty else []) +
            list(expenses_df['property_name'].unique() if not expenses_df.empty else [])
        ))
        
        ws_llc = wb.create_sheet(title="LLC Summary", index=0)
        self._write_llc_summary(ws_llc, income_df, expenses_df, year)
        
        ws_matrix = wb.create_sheet(title="Expense Matrix", index=1)
        self._write_expense_matrix(ws_matrix, expenses_df, all_properties)
        
        for prop in all_properties:
            ws_prop = wb.create_sheet(title=prop[:31])
            self._write_property_detail(ws_prop, prop, income_df, expenses_df)
        
        if output_file is None:
            output_file = os.path.join(self.reports_dir, f"Lust_Rentals_{year}_Complete_Report.xlsx")
        
        wb.save(output_file)
        logger.info(f"Phase 1 report saved to {output_file}")
        return output_file

    def _style_header(self, cell):
        """Apply header styling."""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

    def _style_data(self, cell, bold=False):
        """Apply data cell styling."""
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if bold:
            cell.font = Font(bold=True)

    def _write_llc_summary(self, ws, income_df, expenses_df, year):
        """Write LLC-level summary sheet."""
        ws['A1'] = f"LUST RENTALS LLC - {year} CONSOLIDATED SUMMARY"
        self._style_header(ws['A1'])
        ws.merge_cells('A1:B1')
        ws.row_dimensions[1].height = 25
        
        row = 3
        total_income = income_df['amount'].sum() if not income_df.empty else 0
        total_expenses = expenses_df['amount'].sum() if not expenses_df.empty else 0
        net = total_income - total_expenses
        
        metrics = [
            ("TOTAL INCOME", total_income),
            ("TOTAL EXPENSES", total_expenses),
            ("NET INCOME", net),
            ("EXPENSE RATIO", total_expenses / total_income if total_income > 0 else 0),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws[f'A{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                         top=Side(style='thin'), bottom=Side(style='thin'))
            
            ws[f'B{row}'] = value
            if label == "EXPENSE RATIO":
                ws[f'B{row}'].number_format = '0.0%'
            else:
                ws[f'B{row}'].number_format = '$#,##0.00'
            self._style_data(ws[f'B{row}'], bold=(label == "NET INCOME"))
            row += 1
        
        row += 2
        ws[f'A{row}'] = "EXPENSE BREAKDOWN BY CATEGORY"
        ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Category"
        ws[f'B{row}'] = "Amount"
        self._style_header(ws[f'A{row}'])
        self._style_header(ws[f'B{row}'])
        
        if not expenses_df.empty and 'category_display' in expenses_df.columns:
            for category, amount in expenses_df.groupby('category_display')['amount'].sum().sort_values(ascending=False).items():
                row += 1
                ws[f'A{row}'] = category
                ws[f'B{row}'] = amount
                ws[f'B{row}'].number_format = '$#,##0.00'
                self._style_data(ws[f'B{row}'])
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15

    def _write_expense_matrix(self, ws, expenses_df, all_properties):
        """Write expense matrix sheet (properties × categories)."""
        ws['A1'] = "EXPENSE MATRIX - All Properties × All Categories"
        self._style_header(ws['A1'])
        ws.merge_cells('A1:Z1')
        
        categories = sorted(expenses_df['category_display'].unique()) if not expenses_df.empty else []
        
        row = 3
        ws[f'A{row}'] = "Property"
        self._style_header(ws[f'A{row}'])
        
        for col_idx, cat in enumerate(categories, start=2):
            col = get_column_letter(col_idx)
            ws[f'{col}{row}'] = cat
            self._style_header(ws[f'{col}{row}'])
        
        total_col = get_column_letter(len(categories) + 2)
        ws[f'{total_col}{row}'] = "TOTAL"
        self._style_header(ws[f'{total_col}{row}'])
        
        for prop in all_properties:
            row += 1
            ws[f'A{row}'] = prop
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                         top=Side(style='thin'), bottom=Side(style='thin'))
            
            prop_total = 0
            for col_idx, cat in enumerate(categories, start=2):
                col = get_column_letter(col_idx)
                amount = expenses_df[
                    (expenses_df['property_name'] == prop) & (expenses_df['category_display'] == cat)
                ]['amount'].sum()
                ws[f'{col}{row}'] = amount
                ws[f'{col}{row}'].number_format = '$#,##0.00'
                self._style_data(ws[f'{col}{row}'])
                prop_total += amount
            
            ws[f'{total_col}{row}'] = prop_total
            ws[f'{total_col}{row}'].number_format = '$#,##0.00'
            ws[f'{total_col}{row}'].font = Font(bold=True)
            self._style_data(ws[f'{total_col}{row}'])
        
        row += 1
        ws[f'A{row}'] = "TOTAL"
        self._style_header(ws[f'A{row}'])
        
        for col_idx, cat in enumerate(categories, start=2):
            col = get_column_letter(col_idx)
            total = expenses_df[expenses_df['category_display'] == cat]['amount'].sum()
            ws[f'{col}{row}'] = total
            ws[f'{col}{row}'].number_format = '$#,##0.00'
            self._style_header(ws[f'{col}{row}'])
        
        grand_total = expenses_df['amount'].sum()
        ws[f'{total_col}{row}'] = grand_total
        ws[f'{total_col}{row}'].number_format = '$#,##0.00'
        self._style_header(ws[f'{total_col}{row}'])

    def _write_property_detail(self, ws, property_name, income_df, expenses_df):
        """Write detailed property sheet."""
        ws['A1'] = f"{property_name.upper()} - DETAILED REPORT"
        self._style_header(ws['A1'])
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 25
        
        row = 3
        prop_income_df = income_df[income_df['property_name'] == property_name]
        prop_expenses_df = expenses_df[expenses_df['property_name'] == property_name]
        
        total_income = prop_income_df['amount'].sum() if not prop_income_df.empty else 0
        total_expenses = prop_expenses_df['amount'].sum() if not prop_expenses_df.empty else 0
        net = total_income - total_expenses
        
        # Summary metrics
        metrics = [
            ("TOTAL INCOME", total_income),
            ("TOTAL EXPENSES", total_expenses),
            ("NET INCOME", net),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '$#,##0.00'
            self._style_data(ws[f'B{row}'], bold=(label == "NET INCOME"))
            row += 1
        
        # Income transactions
        row += 2
        ws[f'A{row}'] = "INCOME TRANSACTIONS"
        ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        for col, header in enumerate(['Date', 'Description', 'Category', 'Amount'], start=1):
            cell = ws.cell(row, col)
            cell.value = header
            self._style_header(cell)
        
        if not prop_income_df.empty:
            for _, trans in prop_income_df.iterrows():
                row += 1
                ws[f'A{row}'] = trans.get('transaction_date', '')
                ws[f'B{row}'] = trans.get('description', '')
                ws[f'C{row}'] = trans.get('category', '')
                ws[f'D{row}'] = trans.get('amount', 0)
                ws[f'D{row}'].number_format = '$#,##0.00'
                self._style_data(ws[f'D{row}'])
        
        # Expense transactions
        row += 2
        ws[f'A{row}'] = "EXPENSE TRANSACTIONS"
        ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        for col, header in enumerate(['Date', 'Description', 'Category', 'Amount'], start=1):
            cell = ws.cell(row, col)
            cell.value = header
            self._style_header(cell)
        
        if not prop_expenses_df.empty:
            for _, trans in prop_expenses_df.iterrows():
                row += 1
                ws[f'A{row}'] = trans.get('transaction_date', '')
                ws[f'B{row}'] = trans.get('description', '')
                ws[f'C{row}'] = trans.get('category_display', '')
                ws[f'D{row}'] = trans.get('amount', 0)
                ws[f'D{row}'].number_format = '$#,##0.00'
                self._style_data(ws[f'D{row}'])
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15

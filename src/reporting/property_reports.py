"""
Simplified Property Income and Expense Reports

This module provides simplified reporting functionality that shows only
Income and Expense by each property, available in both PDF and Excel formats.
"""

import os
import logging
import sqlite3
from typing import Dict, List, Tuple, Optional
from datetime import datetime
from fpdf import FPDF as FPDF2
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from pathlib import Path

from src.data_processing.processor import FinancialDataProcessor
from src.utils.config import load_config
from src.utils.date_helpers import safe_format_date
from src.categorization.category_utils import normalize_category, get_display_name

# Use FPDF2 (fpdf2 package) for PDF generation
FPDF = FPDF2

logger = logging.getLogger(__name__)


class PropertyReportGenerator:
    """Generates simplified income and expense reports by property."""

    def __init__(self, data_dir: Path = None):
        """
        Initialize the property report generator.

        Args:
            data_dir: Base directory for data files
        """
        config = load_config()
        self.data_dir = data_dir or config.data_dir
        self.processor = FinancialDataProcessor(data_dir=self.data_dir)
        self.reports_dir = os.path.join(self.data_dir, "reports")
        os.makedirs(self.reports_dir, exist_ok=True)

    def get_property_summary(self, year: int) -> Dict:
        """
        Get income and expense summary by property.

        Args:
            year: Tax year to generate report for

        Returns:
            Dictionary containing property summaries with detailed expense types and income entries
        """
        logger.info(f"Generating property summary for year {year}")

        # Process financial data
        result = self.processor.load_processed_data(year)
        income_df = result["income"]
        expenses_df = result["expenses"]

        # Calculate income by property
        income_by_property = {}
        if not income_df.empty and 'property_name' in income_df.columns:
            income_summary = income_df.groupby('property_name')['amount'].sum()
            income_by_property = income_summary.to_dict()

        # Calculate expenses by property and expense type
        expenses_by_property = {}
        expenses_by_property_and_type = {}
        if not expenses_df.empty and 'property_name' in expenses_df.columns:
            # Total expenses by property
            expense_summary = expenses_df.groupby('property_name')['amount'].sum()
            expenses_by_property = expense_summary.to_dict()

            # Expenses grouped by property and category (expense type)
            if 'category' in expenses_df.columns:
                expense_detail = expenses_df.groupby(['property_name', 'category'])['amount'].sum()
                for (prop, category), amount in expense_detail.items():
                    if prop not in expenses_by_property_and_type:
                        expenses_by_property_and_type[prop] = []
                    expenses_by_property_and_type[prop].append({
                        'type': category,
                        'amount': amount
                    })
                # Sort expense types by amount (descending) for each property
                for prop in expenses_by_property_and_type:
                    expenses_by_property_and_type[prop].sort(key=lambda x: x['amount'], reverse=True)

        # Get income entries by property with dates
        income_entries_by_property = {}
        if not income_df.empty and 'property_name' in income_df.columns:
            # Ensure we have date column
            date_col = None
            for col in income_df.columns:
                if 'date' in col.lower():
                    date_col = col
                    break

            if date_col:
                income_with_dates = income_df[[date_col, 'property_name', 'amount']].copy()
                # Sort by date
                income_with_dates = income_with_dates.sort_values(date_col)

                for prop in income_with_dates['property_name'].unique():
                    prop_income = income_with_dates[income_with_dates['property_name'] == prop]
                    income_entries_by_property[prop] = [
                        {
                            'date': row[date_col],
                            'amount': row['amount']
                        }
                        for _, row in prop_income.iterrows()
                    ]

        # Get expense entries by property with dates and types
        expense_entries_by_property = {}
        if not expenses_df.empty and 'property_name' in expenses_df.columns:
            # Ensure we have date column
            date_col = None
            for col in expenses_df.columns:
                if 'date' in col.lower():
                    date_col = col
                    break

            if date_col and 'category' in expenses_df.columns:
                expenses_with_dates = expenses_df[[date_col, 'property_name', 'category', 'amount']].copy()
                # Sort by date
                expenses_with_dates = expenses_with_dates.sort_values(date_col)

                for prop in expenses_with_dates['property_name'].unique():
                    prop_expenses = expenses_with_dates[expenses_with_dates['property_name'] == prop]
                    expense_entries_by_property[prop] = [
                        {
                            'date': row[date_col],
                            'type': row['category'],
                            'amount': row['amount']
                        }
                        for _, row in prop_expenses.iterrows()
                    ]

        # Get all unique properties
        all_properties = set(list(income_by_property.keys()) + list(expenses_by_property.keys()))

        # Build property summaries
        property_data = []
        total_income = 0
        total_expenses = 0

        for property_name in sorted(all_properties):
            income = income_by_property.get(property_name, 0)
            expenses = expenses_by_property.get(property_name, 0)
            net = income - expenses

            property_data.append({
                'property': property_name,
                'income': income,
                'expenses': expenses,
                'net': net,
                'expense_types': expenses_by_property_and_type.get(property_name, []),
                'income_entries': income_entries_by_property.get(property_name, []),
                'expense_entries': expense_entries_by_property.get(property_name, [])
            })

            total_income += income
            total_expenses += expenses

        return {
            'year': year,
            'properties': property_data,
            'total_income': total_income,
            'total_expenses': total_expenses,
            'total_net': total_income - total_expenses
        }

    def _format_expense_type(self, category: Optional[str]) -> str:
        """Format an expense category for display in the detail sheet."""
        normalized = normalize_category(category)
        if normalized == "hoa":
            return "CONDO FEE"
        return get_display_name(normalized).upper()

    def _parse_expense_date(self, value: Optional[str]) -> Optional[datetime]:
        if not value:
            return None
        if isinstance(value, datetime):
            return value
        try:
            return datetime.fromisoformat(str(value))
        except ValueError:
            return None

    def _load_expense_details(self, year: int) -> List[Dict]:
        """Load expense detail rows with overrides and rule-based fallbacks."""
        db_path = Path(self.processor.processed_db_path)
        if not db_path.exists():
            raise FileNotFoundError("Processed database not found. Run processing first.")

        overrides_db_path = self.data_dir / "overrides" / "overrides.db"
        rules_manager = getattr(self.processor, "rules_manager", None)

        rows: List[Dict] = []
        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            overrides_attached = overrides_db_path.exists()
            if overrides_attached:
                cursor.execute(f"ATTACH DATABASE '{overrides_db_path}' AS overrides_db")

            if overrides_attached:
                query = """
                    SELECT
                        pe.date,
                        pe.debit_amount,
                        pe.description,
                        pe.reference,
                        pe.memo,
                        pe.account_name,
                        pe.code,
                        pe.amount,
                        pe.transaction_id,
                        COALESCE(eo.category, pe.category) as category,
                        COALESCE(eo.property_name, pe.property_name) as property_name
                    FROM processed_expenses pe
                    LEFT JOIN overrides_db.expense_overrides eo ON pe.transaction_id = eo.transaction_id
                    WHERE strftime('%Y', pe.date) = ?
                    ORDER BY pe.date ASC
                """
            else:
                query = """
                    SELECT
                        pe.date,
                        pe.debit_amount,
                        pe.description,
                        pe.reference,
                        pe.memo,
                        pe.account_name,
                        pe.code,
                        pe.amount,
                        pe.transaction_id,
                        pe.category as category,
                        pe.property_name as property_name
                    FROM processed_expenses pe
                    WHERE strftime('%Y', pe.date) = ?
                    ORDER BY pe.date ASC
                """
            cursor.execute(query, (str(year),))
            rows = [dict(row) for row in cursor.fetchall()]

            if overrides_attached:
                cursor.execute("DETACH DATABASE overrides_db")

        details: List[Dict] = []
        for row in rows:
            property_name = row.get("property_name")
            category = row.get("category")

            needs_property = not property_name or str(property_name).strip().upper() == "UNASSIGNED"
            needs_category = not category or str(category).strip() == ""

            if rules_manager and (needs_property or needs_category):
                tx_data = {
                    "description": row.get("description", "") or "",
                    "memo": row.get("memo", "") or "",
                    "amount": str(row.get("amount", 0.0) or 0.0),
                    "payee": row.get("account_name", "") or "",
                }
                actions, _ = rules_manager.evaluate_transaction(tx_data)
                for action in actions:
                    action_type = action.get("type")
                    action_value = action.get("value")
                    if action_type == "set_property" and needs_property and action_value:
                        property_name = action_value
                        needs_property = False
                    elif action_type == "set_category" and needs_category and action_value:
                        category = action_value
                        needs_category = False

            if not property_name or str(property_name).strip() == "":
                property_name = "Unassigned"

            debit_amount = row.get("debit_amount")
            if debit_amount is None:
                debit_amount = row.get("amount")
            if debit_amount is None:
                debit_amount = 0.0
            debit_amount = abs(float(debit_amount))

            details.append({
                "date": self._parse_expense_date(row.get("date")),
                "debit_amount": debit_amount,
                "description": row.get("description", ""),
                "reference": row.get("reference", ""),
                "memo": row.get("memo", ""),
                "vendor": row.get("memo", "") or row.get("description", ""),
                "prop_code": row.get("code", ""),
                "prop_name": property_name,
                "expense_type": self._format_expense_type(category),
            })

        return details

    def _load_expense_pivot_summary(self, year: int) -> List[Dict]:
        """Load grouped expense totals by property and category for pivot summary."""
        db_path = Path(self.processor.processed_db_path)
        if not db_path.exists():
            raise FileNotFoundError("Processed database not found. Run processing first.")

        overrides_db_path = self.data_dir / "overrides" / "overrides.db"

        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            overrides_attached = overrides_db_path.exists()
            if overrides_attached:
                cursor.execute(f"ATTACH DATABASE '{overrides_db_path}' AS overrides_db")
                query = """
                    WITH expense_rows AS (
                        SELECT
                            COALESCE(NULLIF(TRIM(eo.property_name), ''), NULLIF(TRIM(pe.property_name), ''), 'Unassigned') AS property_name,
                            COALESCE(NULLIF(TRIM(eo.category), ''), NULLIF(TRIM(pe.category), ''), 'other') AS category,
                            ABS(COALESCE(pe.debit_amount, pe.amount, 0)) AS debit_amount
                        FROM processed_expenses pe
                        LEFT JOIN overrides_db.expense_overrides eo
                            ON pe.transaction_id = eo.transaction_id
                        WHERE strftime('%Y', pe.date) = ?
                    )
                    SELECT
                        property_name,
                        category,
                        SUM(debit_amount) AS debit_sum,
                        COUNT(*) AS line_count
                    FROM expense_rows
                    GROUP BY property_name, category
                    ORDER BY property_name, debit_sum DESC, category
                """
            else:
                query = """
                    WITH expense_rows AS (
                        SELECT
                            COALESCE(NULLIF(TRIM(pe.property_name), ''), 'Unassigned') AS property_name,
                            COALESCE(NULLIF(TRIM(pe.category), ''), 'other') AS category,
                            ABS(COALESCE(pe.debit_amount, pe.amount, 0)) AS debit_amount
                        FROM processed_expenses pe
                        WHERE strftime('%Y', pe.date) = ?
                    )
                    SELECT
                        property_name,
                        category,
                        SUM(debit_amount) AS debit_sum,
                        COUNT(*) AS line_count
                    FROM expense_rows
                    GROUP BY property_name, category
                    ORDER BY property_name, debit_sum DESC, category
                """

            cursor.execute(query, (str(year),))
            rows = [dict(row) for row in cursor.fetchall()]

            if overrides_attached:
                cursor.execute("DETACH DATABASE overrides_db")

        summary_rows: List[Dict] = []
        for row in rows:
            property_name = row.get("property_name")
            if not property_name or str(property_name).strip() == "":
                property_name = "Unassigned"
            category = row.get("category")

            summary_rows.append({
                "property_name": str(property_name),
                "expense_type": self._format_expense_type(category),
                "debit_amount": float(row.get("debit_sum") or 0.0),
                "line_count": int(row.get("line_count") or 0),
            })

        return summary_rows

    def generate_pdf_report(self, year: int, save_to_file: bool = True) -> Tuple[str, Dict]:
        """
        Generate a PDF report showing detailed expenses by type and income by date for each property.

        Args:
            year: Tax year to generate report for
            save_to_file: Whether to save the PDF to a file

        Returns:
            Tuple of (file_path, summary_data)
        """
        logger.info(f"Generating property PDF report for year {year}")

        # Get property summary data
        summary = self.get_property_summary(year)

        # Create PDF
        pdf = FPDF()
        pdf.add_page()

        # Title
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, 'Lust Rentals', 0, 1, 'C')
        pdf.set_font('Arial', '', 12)
        pdf.cell(0, 10, f'Property Income & Expense Report - {year}', 0, 1, 'C')
        pdf.ln(5)

        # Expenses by Property and Type Section
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, 'EXPENSES BY PROPERTY AND TYPE', 0, 1, 'L')

        # Table header for expenses
        pdf.set_font('Arial', 'B', 10)
        pdf.set_fill_color(37, 99, 235)  # Blue header
        pdf.set_text_color(255, 255, 255)  # White text
        pdf.cell(65, 8, 'Prop Name', 1, 0, 'L', True)
        pdf.cell(70, 8, 'Expense Type', 1, 0, 'L', True)
        pdf.cell(55, 8, 'Debit Amount (Sum)', 1, 1, 'R', True)
        pdf.set_text_color(0, 0, 0)  # Reset to black

        # Property expense details
        pdf.set_font('Arial', '', 9)
        grand_total_expenses = 0
        for prop in summary['properties']:
            if not prop['expense_types']:
                continue

            # First expense type row includes property name
            first_expense = prop['expense_types'][0]
            pdf.cell(65, 7, prop['property'][:30], 1, 0, 'L')
            pdf.cell(70, 7, first_expense['type'][:35], 1, 0, 'L')
            pdf.cell(55, 7, f"${first_expense['amount']:,.2f}", 1, 1, 'R')

            # Remaining expense types (property name blank)
            for expense in prop['expense_types'][1:]:
                pdf.cell(65, 7, '', 1, 0, 'L')
                pdf.cell(70, 7, expense['type'][:35], 1, 0, 'L')
                pdf.cell(55, 7, f"${expense['amount']:,.2f}", 1, 1, 'R')

            # Property total row
            pdf.set_font('Arial', 'B', 9)
            pdf.set_fill_color(229, 231, 235)  # Light gray
            pdf.cell(65, 7, f"{prop['property'][:25]} Total", 1, 0, 'L', True)
            pdf.cell(70, 7, '', 1, 0, 'L', True)
            pdf.cell(55, 7, f"${prop['expenses']:,.2f}", 1, 1, 'R', True)
            pdf.set_font('Arial', '', 9)
            grand_total_expenses += prop['expenses']

        # Grand total for expenses
        pdf.set_font('Arial', 'B', 10)
        pdf.set_fill_color(209, 213, 219)  # Darker gray
        pdf.cell(65, 8, 'GRAND TOTAL', 1, 0, 'L', True)
        pdf.cell(70, 8, '', 1, 0, 'L', True)
        pdf.cell(55, 8, f"${grand_total_expenses:,.2f}", 1, 1, 'R', True)
        pdf.ln(5)

        # Check if we need a new page
        if pdf.get_y() > 200:
            pdf.add_page()

        # Income by Property and Date Section
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, 'INCOME BY PROPERTY AND DATE', 0, 1, 'L')

        # Table header for income
        pdf.set_font('Arial', 'B', 10)
        pdf.set_fill_color(37, 99, 235)  # Blue header
        pdf.set_text_color(255, 255, 255)  # White text
        pdf.cell(65, 8, 'Prop Name', 1, 0, 'L', True)
        pdf.cell(70, 8, 'Deposit Date', 1, 0, 'L', True)
        pdf.cell(55, 8, 'Credit Amount', 1, 1, 'R', True)
        pdf.set_text_color(0, 0, 0)  # Reset to black

        # Property income details
        pdf.set_font('Arial', '', 9)
        grand_total_income = 0
        for prop in summary['properties']:
            if not prop['income_entries']:
                continue

            # Check if we need a new page
            if pdf.get_y() > 250:
                pdf.add_page()
                # Reprint header
                pdf.set_font('Arial', 'B', 10)
                pdf.set_fill_color(37, 99, 235)
                pdf.set_text_color(255, 255, 255)
                pdf.cell(65, 8, 'Prop Name', 1, 0, 'L', True)
                pdf.cell(70, 8, 'Deposit Date', 1, 0, 'L', True)
                pdf.cell(55, 8, 'Credit Amount', 1, 1, 'R', True)
                pdf.set_text_color(0, 0, 0)
                pdf.set_font('Arial', '', 9)

            # First income entry includes property name
            first_income = prop['income_entries'][0]
            date_str = safe_format_date(first_income['date'])
            pdf.cell(65, 7, prop['property'][:30], 1, 0, 'L')
            pdf.cell(70, 7, date_str, 1, 0, 'L')
            pdf.cell(55, 7, f"${first_income['amount']:,.2f}", 1, 1, 'R')

            # Remaining income entries (property name blank)
            for income in prop['income_entries'][1:]:
                if pdf.get_y() > 260:
                    pdf.add_page()
                    # Reprint header
                    pdf.set_font('Arial', 'B', 10)
                    pdf.set_fill_color(37, 99, 235)
                    pdf.set_text_color(255, 255, 255)
                    pdf.cell(65, 8, 'Prop Name', 1, 0, 'L', True)
                    pdf.cell(70, 8, 'Deposit Date', 1, 0, 'L', True)
                    pdf.cell(55, 8, 'Credit Amount', 1, 1, 'R', True)
                    pdf.set_text_color(0, 0, 0)
                    pdf.set_font('Arial', '', 9)

                date_str = safe_format_date(income['date'])
                pdf.cell(65, 7, '', 1, 0, 'L')
                pdf.cell(70, 7, date_str, 1, 0, 'L')
                pdf.cell(55, 7, f"${income['amount']:,.2f}", 1, 1, 'R')

            # Property total row
            pdf.set_font('Arial', 'B', 9)
            pdf.set_fill_color(229, 231, 235)  # Light gray
            pdf.cell(65, 7, f"{prop['property'][:25]} Total", 1, 0, 'L', True)
            pdf.cell(70, 7, '', 1, 0, 'L', True)
            pdf.cell(55, 7, f"${prop['income']:,.2f}", 1, 1, 'R', True)
            pdf.set_font('Arial', '', 9)
            grand_total_income += prop['income']

        # Grand total for income
        pdf.set_font('Arial', 'B', 10)
        pdf.set_fill_color(209, 213, 219)  # Darker gray
        pdf.cell(65, 8, 'GRAND TOTAL', 1, 0, 'L', True)
        pdf.cell(70, 8, '', 1, 0, 'L', True)
        pdf.cell(55, 8, f"${grand_total_income:,.2f}", 1, 1, 'R', True)
        pdf.ln(10)

        # ========== NEW SECTION: Expenses by Property, Date and Type ==========
        pdf.add_page()
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, 'EXPENSES BY PROPERTY, DATE AND TYPE', 0, 1, 'L')

        # Table header for expense entries
        pdf.set_font('Arial', 'B', 10)
        pdf.set_fill_color(37, 99, 235)  # Blue header
        pdf.set_text_color(255, 255, 255)  # White text
        pdf.cell(50, 8, 'Prop Name', 1, 0, 'L', True)
        pdf.cell(45, 8, 'Expense Date', 1, 0, 'L', True)
        pdf.cell(50, 8, 'Expense Type', 1, 0, 'L', True)
        pdf.cell(45, 8, 'Debit Amount', 1, 1, 'R', True)
        pdf.set_text_color(0, 0, 0)  # Reset to black

        # Property expense entry details
        pdf.set_font('Arial', '', 9)
        grand_total_expenses_detailed = 0
        for prop in summary['properties']:
            if not prop['expense_entries']:
                continue

            # Check if we need a new page
            if pdf.get_y() > 250:
                pdf.add_page()
                # Reprint header
                pdf.set_font('Arial', 'B', 10)
                pdf.set_fill_color(37, 99, 235)
                pdf.set_text_color(255, 255, 255)
                pdf.cell(50, 8, 'Prop Name', 1, 0, 'L', True)
                pdf.cell(45, 8, 'Expense Date', 1, 0, 'L', True)
                pdf.cell(50, 8, 'Expense Type', 1, 0, 'L', True)
                pdf.cell(45, 8, 'Debit Amount', 1, 1, 'R', True)
                pdf.set_text_color(0, 0, 0)
                pdf.set_font('Arial', '', 9)

            # First expense entry row includes property name
            first_expense = prop['expense_entries'][0]
            date_str = safe_format_date(first_expense['date'])
            pdf.cell(50, 7, prop['property'][:25], 1, 0, 'L')
            pdf.cell(45, 7, date_str, 1, 0, 'L')
            pdf.cell(50, 7, first_expense['type'][:25], 1, 0, 'L')
            pdf.cell(45, 7, f"${first_expense['amount']:,.2f}", 1, 1, 'R')

            # Remaining expense entries (property name blank)
            for expense in prop['expense_entries'][1:]:
                if pdf.get_y() > 260:
                    pdf.add_page()
                    # Reprint header
                    pdf.set_font('Arial', 'B', 10)
                    pdf.set_fill_color(37, 99, 235)
                    pdf.set_text_color(255, 255, 255)
                    pdf.cell(50, 8, 'Prop Name', 1, 0, 'L', True)
                    pdf.cell(45, 8, 'Expense Date', 1, 0, 'L', True)
                    pdf.cell(50, 8, 'Expense Type', 1, 0, 'L', True)
                    pdf.cell(45, 8, 'Debit Amount', 1, 1, 'R', True)
                    pdf.set_text_color(0, 0, 0)
                    pdf.set_font('Arial', '', 9)

                date_str = safe_format_date(expense['date'])
                pdf.cell(50, 7, '', 1, 0, 'L')
                pdf.cell(45, 7, date_str, 1, 0, 'L')
                pdf.cell(50, 7, expense['type'][:25], 1, 0, 'L')
                pdf.cell(45, 7, f"${expense['amount']:,.2f}", 1, 1, 'R')

            # Property total row
            pdf.set_font('Arial', 'B', 9)
            pdf.set_fill_color(229, 231, 235)  # Light gray
            pdf.cell(50, 7, f"{prop['property'][:20]} Total", 1, 0, 'L', True)
            pdf.cell(45, 7, '', 1, 0, 'L', True)
            pdf.cell(50, 7, '', 1, 0, 'L', True)
            pdf.cell(45, 7, f"${prop['expenses']:,.2f}", 1, 1, 'R', True)
            pdf.set_font('Arial', '', 9)
            grand_total_expenses_detailed += prop['expenses']

        # Grand total for expenses
        pdf.set_font('Arial', 'B', 10)
        pdf.set_fill_color(209, 213, 219)  # Darker gray
        pdf.cell(50, 8, 'GRAND TOTAL', 1, 0, 'L', True)
        pdf.cell(45, 8, '', 1, 0, 'L', True)
        pdf.cell(50, 8, '', 1, 0, 'L', True)
        pdf.cell(45, 8, f"${grand_total_expenses_detailed:,.2f}", 1, 1, 'R', True)
        pdf.ln(10)

        # Overall Summary
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, 'OVERALL SUMMARY', 0, 1, 'L')
        pdf.set_font('Arial', '', 10)

        pdf.cell(80, 8, 'Total Income:', 0, 0)
        pdf.cell(0, 8, f"${summary['total_income']:,.2f}", 0, 1)

        pdf.cell(80, 8, 'Total Expenses:', 0, 0)
        pdf.cell(0, 8, f"${summary['total_expenses']:,.2f}", 0, 1)

        pdf.cell(80, 8, 'Net Income:', 0, 0)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 8, f"${summary['total_net']:,.2f}", 0, 1)

        # Footer
        pdf.ln(10)
        pdf.set_font('Arial', 'I', 8)
        pdf.cell(0, 5, f'Report generated on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 0, 1, 'C')

        # Save file
        file_path = None
        if save_to_file:
            file_path = os.path.join(self.reports_dir, f'property_report_{year}.pdf')
            pdf.output(file_path)
            logger.info(f"Property PDF report saved to: {file_path}")

        return file_path, summary

    def generate_excel_report(self, year: int, save_to_file: bool = True) -> Tuple[str, Dict]:
        """
        Generate an Excel report showing yearly income and expenses with property breakdown.

        Args:
            year: Tax year to generate report for
            save_to_file: Whether to save the Excel file

        Returns:
            Tuple of (file_path, summary_data)
        """
        logger.info(f"Generating yearly income/expense Excel report for year {year}")

        db_path = Path(self.processor.processed_db_path)
        if not db_path.exists():
            raise FileNotFoundError("Processed database not found. Run processing first.")

        try:
            with sqlite3.connect(db_path) as conn:
                income_df = pd.read_sql_query("SELECT * FROM processed_income", conn)
                expenses_df = pd.read_sql_query("SELECT * FROM processed_expenses", conn)
        except sqlite3.OperationalError as exc:
            if "no such table" in str(exc):
                raise FileNotFoundError(
                    "No processed data found. Please process your bank transactions first using the 'Run processor' button."
                ) from exc
            raise
        except sqlite3.Error as exc:
            raise RuntimeError(f"Failed to read processed data: {exc}") from exc

        if income_df.empty and expenses_df.empty:
            raise FileNotFoundError("No processed data found for the requested year.")

        if 'date' in income_df.columns:
            income_df['date'] = pd.to_datetime(income_df['date'], errors='coerce')
            income_df = income_df[income_df['date'].dt.year == year]

        if 'date' in expenses_df.columns:
            expenses_df['date'] = pd.to_datetime(expenses_df['date'], errors='coerce')
            expenses_df = expenses_df[expenses_df['date'].dt.year == year]

        if 'amount' in income_df.columns:
            income_df['amount'] = income_df['amount'].abs()

        if 'amount' in expenses_df.columns:
            expenses_df['amount'] = expenses_df['amount'].abs()

        if income_df.empty and expenses_df.empty:
            raise FileNotFoundError("No processed data found for the requested year.")

        property_breakdown = []
        if 'property_name' in income_df.columns or 'property_name' in expenses_df.columns:
            income_by_property = (
                income_df.groupby('property_name')['amount'].sum()
                if not income_df.empty and 'property_name' in income_df.columns and 'amount' in income_df.columns
                else pd.Series(dtype=float)
            )
            expenses_by_property = (
                expenses_df.groupby('property_name')['amount'].sum()
                if not expenses_df.empty and 'property_name' in expenses_df.columns and 'amount' in expenses_df.columns
                else pd.Series(dtype=float)
            )

            properties = sorted(set(income_by_property.index.tolist()) | set(expenses_by_property.index.tolist()))
            for prop in properties:
                if prop is None or str(prop).strip() == "":
                    prop_name = "Unassigned"
                else:
                    prop_name = str(prop)
                income_total = float(income_by_property.get(prop, 0.0))
                expense_total = float(expenses_by_property.get(prop, 0.0))
                property_breakdown.append({
                    "Property": prop_name,
                    "Income": round(income_total, 2),
                    "Expenses": round(expense_total, 2),
                })

        property_breakdown_df = pd.DataFrame(property_breakdown)
        if not property_breakdown_df.empty:
            property_breakdown_df = property_breakdown_df.sort_values("Property")

        total_income = float(income_df['amount'].sum()) if 'amount' in income_df.columns else 0.0
        total_expenses = float(expenses_df['amount'].sum()) if 'amount' in expenses_df.columns else 0.0
        total_net = total_income - total_expenses

        summary = {
            "tax_year": year,
            "total_income": total_income,
            "total_expenses": total_expenses,
            "total_net": total_net,
            "property_count": len(property_breakdown_df),
        }

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_income = wb.create_sheet("Income")
        ws_expenses = wb.create_sheet("Expenses")
        ws_properties = wb.create_sheet("Property Breakdown")
        ws_pivot_summary = wb.create_sheet("Pivot Summary")
        ws_expense_details = wb.create_sheet("Expense Details")

        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        title_font = Font(name='Calibri', size=16, bold=True, color='1E40AF')
        subtitle_font = Font(name='Calibri', size=11, color='334155')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        summary_header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
        property_header_fill = PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid')
        detail_header_fill = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')
        pivot_header_fill = PatternFill(start_color='0284C7', end_color='0284C7', fill_type='solid')
        alt_row_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        positive_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
        negative_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1'),
        )

        # Summary sheet
        ws_summary['A1'] = 'Yearly Income & Expense - Lust Rentals LLC'
        ws_summary['A1'].font = title_font
        ws_summary['A2'] = f'Tax Year: {year}'
        ws_summary['A2'].font = subtitle_font
        ws_summary.merge_cells('A1:B1')

        ws_summary['A4'] = 'Metric'
        ws_summary['B4'] = 'Value'
        for col in range(1, 3):
            cell = ws_summary.cell(row=4, column=col)
            cell.font = header_font
            cell.fill = summary_header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        amount_col_income = None
        for idx, col in enumerate(income_df.columns, start=1):
            if 'amount' in col.lower():
                amount_col_income = idx
                break
        amount_col_expense = None
        for idx, col in enumerate(expenses_df.columns, start=1):
            if 'amount' in col.lower():
                amount_col_expense = idx
                break

        income_amount_col = get_column_letter(amount_col_income) if amount_col_income else None
        expense_amount_col = get_column_letter(amount_col_expense) if amount_col_expense else None

        summary_rows = [
            ('Tax Year', year),
            ('Total Income', f"=SUM(Income!{income_amount_col}:{income_amount_col})" if income_amount_col else total_income),
            ('Total Expenses', f"=SUM(Expenses!{expense_amount_col}:{expense_amount_col})" if expense_amount_col else total_expenses),
            ('Net Income', '=B6-B7'),
            ('Income Transactions', '=MAX(COUNTA(Income!A:A)-1,0)'),
            ('Expense Transactions', '=MAX(COUNTA(Expenses!A:A)-1,0)'),
        ]

        start_row = 5
        for offset, (label, value) in enumerate(summary_rows):
            row = start_row + offset
            ws_summary.cell(row=row, column=1, value=label)
            ws_summary.cell(row=row, column=2, value=value)

            for col in range(1, 3):
                cell = ws_summary.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left' if col == 1 else 'right')
                cell.fill = alt_row_fill if row % 2 == 0 else white_fill
                if col == 1:
                    cell.font = Font(name='Calibri', size=11, bold=True, color='1E293B')
                else:
                    cell.font = Font(name='Calibri', size=11, color='334155')

        for row in [6, 7, 8]:
            ws_summary.cell(row=row, column=2).number_format = '$#,##0.00'

        ws_summary.conditional_formatting.add(
            'B8',
            CellIsRule(operator='greaterThanOrEqual', formula=['0'], fill=positive_fill)
        )
        ws_summary.conditional_formatting.add(
            'B8',
            CellIsRule(operator='lessThan', formula=['0'], fill=negative_fill)
        )

        ws_summary['A12'] = f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary['A12'].font = Font(name='Calibri', size=9, italic=True, color='64748B')

        # Income sheet
        if not income_df.empty or len(income_df.columns) > 0:
            for row in dataframe_to_rows(income_df, index=False, header=True):
                ws_income.append(row)
            if income_df.empty:
                ws_income.append(["No income data available for this year."])
        else:
            ws_income.append(["No income data available for this year."])

        # Expenses sheet
        if not expenses_df.empty or len(expenses_df.columns) > 0:
            for row in dataframe_to_rows(expenses_df, index=False, header=True):
                ws_expenses.append(row)
            if expenses_df.empty:
                ws_expenses.append(["No expense data available for this year."])
        else:
            ws_expenses.append(["No expense data available for this year."])

        # Property breakdown sheet
        if not property_breakdown_df.empty:
            ws_properties.append(["Property", "Income", "Expenses", "Net"])
            for idx, row in property_breakdown_df.iterrows():
                excel_row = ws_properties.max_row + 1
                ws_properties.cell(row=excel_row, column=1, value=row["Property"])
                ws_properties.cell(row=excel_row, column=2, value=row["Income"])
                ws_properties.cell(row=excel_row, column=3, value=row["Expenses"])
                ws_properties.cell(row=excel_row, column=4, value=f"=B{excel_row}-C{excel_row}")

            total_row = ws_properties.max_row + 1
            ws_properties.cell(row=total_row, column=1, value="TOTAL")
            ws_properties.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row-1})")
            ws_properties.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row-1})")
            ws_properties.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row-1})")
        else:
            ws_properties.append(["No property breakdown available for this year."])

        # Expense details sheet
        expense_details = self._load_expense_details(year)
        detail_headers = [
            "Date",
            "Debit Amount",
            "Description",
            "Reference",
            "Memo",
            "Vendor",
            "Prop Code",
            "Prop Name",
            "Expense Type",
        ]

        if expense_details:
            ws_expense_details.append(detail_headers)

            property_groups: Dict[str, Dict[str, List[Dict]]] = {}
            for entry in expense_details:
                prop_name = entry["prop_name"]
                expense_type = entry["expense_type"]
                property_groups.setdefault(prop_name, {}).setdefault(expense_type, []).append(entry)

            current_row = ws_expense_details.max_row + 1
            grand_total = 0.0
            grand_count = 0

            for prop_name in sorted(property_groups.keys(), key=lambda x: x.lower()):
                # Property header row
                ws_expense_details.cell(row=current_row, column=8, value=prop_name)
                for col_idx in range(1, len(detail_headers) + 1):
                    cell = ws_expense_details.cell(row=current_row, column=col_idx)
                    cell.font = Font(name='Calibri', size=11, bold=True)
                    cell.fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
                    cell.border = thin_border
                current_row += 1

                for expense_type in sorted(property_groups[prop_name].keys()):
                    entries = property_groups[prop_name][expense_type]
                    subtotal = sum(e["debit_amount"] for e in entries)
                    count = len(entries)

                    subtotal_label = f"{expense_type} +{count} lines,{subtotal:,.2f}"
                    subtotal_row = current_row
                    ws_expense_details.cell(row=subtotal_row, column=9, value=subtotal_label)
                    for col_idx in range(1, len(detail_headers) + 1):
                        cell = ws_expense_details.cell(row=subtotal_row, column=col_idx)
                        cell.font = Font(name='Calibri', size=10, bold=True)
                        cell.border = thin_border
                    current_row += 1

                    detail_start = current_row
                    for entry in entries:
                        ws_expense_details.cell(row=current_row, column=1, value=entry["date"])
                        ws_expense_details.cell(row=current_row, column=2, value=entry["debit_amount"])
                        ws_expense_details.cell(row=current_row, column=3, value=entry["description"])
                        ws_expense_details.cell(row=current_row, column=4, value=entry["reference"])
                        ws_expense_details.cell(row=current_row, column=5, value=entry["memo"])
                        ws_expense_details.cell(row=current_row, column=6, value=entry["vendor"])
                        ws_expense_details.cell(row=current_row, column=7, value=entry["prop_code"])
                        ws_expense_details.cell(row=current_row, column=8, value=entry["prop_name"])
                        ws_expense_details.cell(row=current_row, column=9, value=entry["expense_type"])

                        for col_idx in range(1, len(detail_headers) + 1):
                            cell = ws_expense_details.cell(row=current_row, column=col_idx)
                            cell.border = thin_border
                            if col_idx == 1:
                                cell.number_format = 'yyyy-mm-dd'
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            elif col_idx == 2:
                                cell.number_format = '$#,##0.00'
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            else:
                                cell.alignment = Alignment(horizontal='left', vertical='center')

                        current_row += 1

                    detail_end = current_row - 1
                    if detail_end >= detail_start:
                        ws_expense_details.row_dimensions.group(
                            detail_start, detail_end, outline_level=2, hidden=False
                        )
                    ws_expense_details.row_dimensions.group(
                        subtotal_row, subtotal_row, outline_level=1, hidden=False
                    )

                    grand_total += subtotal
                    grand_count += count

                current_row += 1

            grand_label = f"GRAND TOTAL +{grand_count} lines,{grand_total:,.2f}"
            ws_expense_details.cell(row=current_row, column=9, value=grand_label)
            for col_idx in range(1, len(detail_headers) + 1):
                cell = ws_expense_details.cell(row=current_row, column=col_idx)
                cell.font = Font(name='Calibri', size=11, bold=True)
                cell.fill = PatternFill(start_color='CBD5E1', end_color='CBD5E1', fill_type='solid')
                cell.border = thin_border
        else:
            ws_expense_details.append(["No expense details available for this year."])

        # Pivot summary sheet
        pivot_rows = self._load_expense_pivot_summary(year)
        pivot_headers = ["Prop Name", "Expense Type", "Debit Amount (Sum)"]

        pivot_has_header = False
        if pivot_rows:
            ws_pivot_summary.append(pivot_headers)
            pivot_has_header = True

            grouped: Dict[str, List[Dict]] = {}
            for row in pivot_rows:
                grouped.setdefault(row["property_name"], []).append(row)

            current_row = ws_pivot_summary.max_row + 1
            grand_total = 0.0
            grand_count = 0
            zebra_index = 0

            for prop_name in sorted(grouped.keys(), key=lambda x: x.lower()):
                prop_rows = sorted(grouped[prop_name], key=lambda x: x["debit_amount"], reverse=True)

                ws_pivot_summary.cell(row=current_row, column=1, value=prop_name)
                for col_idx in range(1, 4):
                    cell = ws_pivot_summary.cell(row=current_row, column=col_idx)
                    cell.font = Font(name='Calibri', size=11, bold=True)
                    cell.border = thin_border
                    cell.fill = white_fill
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                current_row += 1

                for row in prop_rows:
                    subtotal_label = f"Sum: {row['expense_type']} +{row['line_count']} lines"
                    ws_pivot_summary.cell(row=current_row, column=2, value=subtotal_label)
                    ws_pivot_summary.cell(row=current_row, column=3, value=row["debit_amount"])

                    row_fill = alt_row_fill if zebra_index % 2 == 0 else white_fill
                    for col_idx in range(1, 4):
                        cell = ws_pivot_summary.cell(row=current_row, column=col_idx)
                        cell.border = thin_border
                        cell.fill = row_fill
                        cell.font = Font(name='Calibri', size=11)
                        if col_idx == 3:
                            cell.number_format = '$#,##0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')

                    current_row += 1
                    zebra_index += 1
                    grand_total += row["debit_amount"]
                    grand_count += row["line_count"]

            grand_label = f"GRAND TOTAL +{grand_count} lines"
            ws_pivot_summary.cell(row=current_row, column=1, value=grand_label)
            ws_pivot_summary.cell(row=current_row, column=3, value=grand_total)
            for col_idx in range(1, 4):
                cell = ws_pivot_summary.cell(row=current_row, column=col_idx)
                cell.font = Font(name='Calibri', size=11, bold=True)
                cell.border = thin_border
                cell.fill = PatternFill(start_color='CBD5E1', end_color='CBD5E1', fill_type='solid')
                if col_idx == 3:
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            ws_pivot_summary.append(["No pivot summary available for this year."])

        # Style income and expense sheets
        for ws in (ws_income, ws_expenses):
            if ws.max_row < 2:
                continue
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            for row_idx in range(2, ws.max_row + 1):
                row_fill = alt_row_fill if row_idx % 2 == 0 else white_fill
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.fill = row_fill
                    column_name = str(ws.cell(row=1, column=col_idx).value or "").lower()
                    if 'amount' in column_name:
                        cell.number_format = '$#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    elif 'date' in column_name:
                        cell.number_format = 'yyyy-mm-dd'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

        # Style property breakdown sheet
        if ws_properties.max_row >= 2 and ws_properties.max_column >= 4:
            for col_idx in range(1, ws_properties.max_column + 1):
                cell = ws_properties.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = property_header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            for row_idx in range(2, ws_properties.max_row + 1):
                row_fill = alt_row_fill if row_idx % 2 == 0 else white_fill
                for col_idx in range(1, ws_properties.max_column + 1):
                    cell = ws_properties.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.fill = row_fill
                    if col_idx in (2, 3, 4):
                        cell.number_format = '$#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

            total_row = ws_properties.max_row
            for col_idx in range(1, ws_properties.max_column + 1):
                cell = ws_properties.cell(row=total_row, column=col_idx)
                cell.font = Font(name='Calibri', size=11, bold=True)
                cell.fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')

        # Style expense details sheet
        if ws_expense_details.max_row >= 1 and ws_expense_details.max_column >= 1:
            header_row = 1
            for col_idx in range(1, ws_expense_details.max_column + 1):
                cell = ws_expense_details.cell(row=header_row, column=col_idx)
                cell.font = header_font
                cell.fill = detail_header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            ws_expense_details.freeze_panes = 'A2'
            ws_expense_details.sheet_properties.outlinePr.summaryBelow = True

        # Style pivot summary sheet
        if pivot_has_header and ws_pivot_summary.max_row >= 1 and ws_pivot_summary.max_column >= 1:
            header_row = 1
            for col_idx in range(1, ws_pivot_summary.max_column + 1):
                cell = ws_pivot_summary.cell(row=header_row, column=col_idx)
                cell.font = header_font
                cell.fill = pivot_header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            ws_pivot_summary.freeze_panes = 'A2'

        # Chart
        if ws_properties.max_row >= 3 and ws_properties.max_column >= 3:
            chart = BarChart()
            chart.title = "Income vs Expenses by Property"
            chart.y_axis.title = "Amount"
            data = Reference(ws_properties, min_col=2, max_col=3, min_row=1, max_row=ws_properties.max_row - 1)
            categories = Reference(ws_properties, min_col=1, min_row=2, max_row=ws_properties.max_row - 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 8
            chart.width = 16
            ws_summary.add_chart(chart, "D4")

        for ws in (ws_summary, ws_income, ws_expenses, ws_properties, ws_pivot_summary, ws_expense_details):
            for column in ws.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                if column_letter:
                    ws.column_dimensions[column_letter].width = min(max(max_length + 3, 12), 50)

            if ws.title in ("Income", "Expenses", "Property Breakdown", "Pivot Summary", "Expense Details"):
                ws.freeze_panes = 'A2'

        file_path = None
        if save_to_file:
            file_path = os.path.join(self.reports_dir, 'Yearly Income & Expense Lust Rentals LLC.xlsx')
            wb.save(file_path)
            logger.info(f"Property Excel report saved to: {file_path}")

        return file_path, summary


def main():
    """Test the property report generator."""
    import sys

    year = int(sys.argv[1]) if len(sys.argv) > 1 else 2025

    generator = PropertyReportGenerator()

    print(f"\nGenerating property reports for {year}...")

    # Generate PDF
    pdf_path, pdf_summary = generator.generate_pdf_report(year)
    print(f"✓ PDF report generated: {pdf_path}")

    # Generate Excel
    excel_path, excel_summary = generator.generate_excel_report(year)
    print(f"✓ Excel report generated: {excel_path}")

    print(f"\nSummary:")
    print(f"  Total Income: ${pdf_summary['total_income']:,.2f}")
    print(f"  Total Expenses: ${pdf_summary['total_expenses']:,.2f}")
    print(f"  Net Income: ${pdf_summary['total_net']:,.2f}")
    print(f"  Properties: {len(pdf_summary['properties'])}")


if __name__ == '__main__':
    main()

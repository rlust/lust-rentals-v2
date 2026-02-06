"""Data export routes for CSV and Excel downloads."""
from __future__ import annotations

import io
import logging
import sqlite3
from typing import Optional

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse
import pandas as pd

from src.api.dependencies import get_config
from src.categorization.category_utils import normalize_category, get_display_name

router = APIRouter()
logger = logging.getLogger(__name__)


def resolve_report_year(year: Optional[int]) -> int:
    """Resolve the report year, defaulting to previous year if not specified."""
    from datetime import datetime
    from src.api.dependencies import get_tax_reporter
    reporter = get_tax_reporter()
    return year or (reporter.current_year - 1)


@router.get("/{dataset}")
def export_dataset(http_request: Request, dataset: str) -> StreamingResponse:
    """Export processed datasets (income/expenses) as CSV for audit."""

    dataset_map = {
        "income": "processed_income",
        "expenses": "processed_expenses",
    }
    table_name = dataset_map.get(dataset.lower())
    if table_name is None:
        raise HTTPException(status_code=404, detail="Dataset not found.")

    db_path = get_config().data_dir / "processed" / "processed.db"
    if not db_path.exists():
        raise HTTPException(status_code=404, detail="Processed database not found. Run processing first.")

    try:
        with sqlite3.connect(db_path) as conn:
            df = pd.read_sql_query(f"SELECT * FROM {table_name}", conn)
    except sqlite3.Error as exc:
        raise HTTPException(status_code=500, detail=f"Failed to read dataset: {exc}") from exc

    if df.empty:
        raise HTTPException(status_code=404, detail="Dataset is empty.")

    buffer = io.StringIO()
    df.to_csv(buffer, index=False)
    buffer.seek(0)

    filename = f"{table_name}.csv"
    return StreamingResponse(
        buffer,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/excel/report")
def export_excel_report(http_request: Request, year: Optional[int] = None) -> StreamingResponse:
    """
    Export comprehensive Excel report with multiple sheets containing:
    - Summary: Annual totals and key metrics
    - Income: Detailed income transactions
    - Expenses: Detailed expense transactions
    - Property Summary: Breakdown by property

    NOTE: This function is 500+ lines and should be refactored into smaller helper functions.
    See improvement #3 in the roadmap: "Extract duplicated Excel styling code"
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    resolved_year = resolve_report_year(year)
    db_path = get_config().data_dir / "processed" / "processed.db"

    if not db_path.exists():
        raise HTTPException(
            status_code=404,
            detail="Processed database not found. Run processing first."
        )

    try:
        # Load data from database
        with sqlite3.connect(db_path) as conn:
            income_df = pd.read_sql_query("SELECT * FROM processed_income", conn)
            expenses_df = pd.read_sql_query("SELECT * FROM processed_expenses", conn)
    except (sqlite3.OperationalError, pd.errors.DatabaseError) as exc:
        if "no such table" in str(exc):
            raise HTTPException(
                status_code=404,
                detail="No processed data found. Please process your bank transactions first using the 'Run processor' button."
            ) from exc
        raise HTTPException(
            status_code=500,
            detail=f"Database error: {exc}"
        ) from exc
    except sqlite3.Error as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to read data: {exc}"
        ) from exc

    if income_df.empty and expenses_df.empty:
        raise HTTPException(
            status_code=404,
            detail="No data available. Process transactions first."
        )

    # Filter by year if date column exists
    if 'date' in income_df.columns:
        income_df['date'] = pd.to_datetime(income_df['date'], errors='coerce')
        income_df = income_df[income_df['date'].dt.year == resolved_year]

    if 'date' in expenses_df.columns:
        expenses_df['date'] = pd.to_datetime(expenses_df['date'], errors='coerce')
        expenses_df = expenses_df[expenses_df['date'].dt.year == resolved_year]

    # Calculate summary metrics
    total_income = income_df['amount'].sum() if not income_df.empty else 0
    total_expenses = expenses_df['amount'].sum() if not expenses_df.empty else 0
    net_income = total_income - abs(total_expenses)

    # Normalize categories in expenses dataframe
    if not expenses_df.empty and 'category' in expenses_df.columns:
        expenses_df['category_normalized'] = expenses_df['category'].apply(normalize_category)
        expenses_df['category_display'] = expenses_df['category_normalized'].apply(get_display_name)

    # Calculate expenses by category
    expense_by_category = []
    if not expenses_df.empty and 'category_normalized' in expenses_df.columns:
        category_summary = expenses_df.groupby(['category_normalized', 'category_display']).agg({
            'amount': ['sum', 'count']
        }).round(2)
        category_summary.columns = ['Total', 'Count']
        category_summary = category_summary.sort_values('Total', ascending=False)

        for (category_norm, category_display), row in category_summary.iterrows():
            expense_by_category.append({
                'Category': category_display,
                'Total': abs(row['Total']),
                'Transaction Count': int(row['Count'])
            })

    expense_by_category_df = pd.DataFrame(expense_by_category) if expense_by_category else pd.DataFrame()

    # Create summary DataFrame with raw values
    summary_data = {
        'Metric': [
            'Tax Year',
            'Total Income',
            'Total Expenses',
            'Net Income',
            'Income Transactions',
            'Expense Transactions',
            '',  # Blank row
            'EXPENSE BREAKDOWN BY CATEGORY',
        ],
        'Value': [
            resolved_year,
            total_income,
            abs(total_expenses),
            net_income,
            len(income_df),
            len(expenses_df),
            '',  # Blank row
            '',  # No value for header
        ]
    }

    # Add expense categories to summary
    if expense_by_category:
        for item in expense_by_category:
            summary_data['Metric'].append(f"  {item['Category']}")
            summary_data['Value'].append(item['Total'])

    summary_df = pd.DataFrame(summary_data)

    # Property breakdown
    property_summary_data = []
    if not income_df.empty and 'property_name' in income_df.columns:
        income_by_property = income_df.groupby('property_name')['amount'].sum()
        for prop, amount in income_by_property.items():
            property_summary_data.append({
                'Property': prop,
                'Income': amount,
                'Income Transactions': len(income_df[income_df['property_name'] == prop])
            })

    if not expenses_df.empty and 'property_name' in expenses_df.columns:
        expense_by_property = expenses_df.groupby('property_name')['amount'].sum()
        for prop, amount in expense_by_property.items():
            existing = next((item for item in property_summary_data if item['Property'] == prop), None)
            if existing:
                existing['Expenses'] = abs(amount)
                existing['Expense Transactions'] = len(expenses_df[expenses_df['property_name'] == prop])
            else:
                property_summary_data.append({
                    'Property': prop,
                    'Income': 0,
                    'Income Transactions': 0,
                    'Expenses': abs(amount),
                    'Expense Transactions': len(expenses_df[expenses_df['property_name'] == prop])
                })

            # Add top 3 expense categories for this property
            if 'category_display' in expenses_df.columns:
                property_expenses = expenses_df[expenses_df['property_name'] == prop]
                top_categories = property_expenses.groupby('category_display')['amount'].sum().abs().sort_values(ascending=False).head(3)
                category_list = []
                for cat, amt in top_categories.items():
                    cat_name = cat if cat and str(cat).strip() else 'Uncategorized'
                    category_list.append(f"{cat_name}: ${amt:,.2f}")

                if existing:
                    existing['Top Categories'] = '; '.join(category_list)
                else:
                    # This shouldn't happen since we just appended, but for safety
                    for item in property_summary_data:
                        if item['Property'] == prop:
                            item['Top Categories'] = '; '.join(category_list)
                            break

    property_summary_df = pd.DataFrame(property_summary_data) if property_summary_data else pd.DataFrame()
    if not property_summary_df.empty:
        property_summary_df['Net'] = property_summary_df.get('Income', 0) - property_summary_df.get('Expenses', 0)

    # Create detailed property expense breakdown by category
    property_expense_breakdown_data = []
    if not expenses_df.empty and 'property_name' in expenses_df.columns and 'category_display' in expenses_df.columns:
        # Get all unique properties
        properties = expenses_df['property_name'].dropna().unique()

        for prop in sorted(properties):
            if not prop or str(prop).strip() == '':
                continue

            # Get expenses for this property grouped by category
            property_expenses = expenses_df[expenses_df['property_name'] == prop]
            category_breakdown = property_expenses.groupby('category_display').agg({
                'amount': ['sum', 'count']
            }).round(2)
            category_breakdown.columns = ['Total', 'Count']
            category_breakdown = category_breakdown.sort_values('Total', ascending=False)

            # Add each category as a row
            for category, row in category_breakdown.iterrows():
                property_expense_breakdown_data.append({
                    'Property': prop,
                    'Category': category,
                    'Amount': abs(row['Total']),
                    'Transaction Count': int(row['Count'])
                })

    property_expense_breakdown_df = pd.DataFrame(property_expense_breakdown_data) if property_expense_breakdown_data else pd.DataFrame()

    # Define styling
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    summary_header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
    property_header_fill = PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid')

    cell_alignment = Alignment(horizontal='left', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    alt_row_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    metric_label_font = Font(name='Calibri', size=11, bold=True, color='1E293B')
    metric_value_font = Font(name='Calibri', size=11, color='334155')

    positive_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    negative_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    # Create Excel file in memory
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Write summary sheet
        summary_df.to_excel(writer, sheet_name='Summary', index=False, startrow=1)
        ws_summary = writer.sheets['Summary']

        # Add title
        ws_summary['A1'] = f'Tax Report Summary - {resolved_year}'
        ws_summary['A1'].font = Font(name='Calibri', size=16, bold=True, color='1E40AF')
        ws_summary['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws_summary.merge_cells('A1:B1')

        # Style summary headers
        for col in range(1, 3):
            cell = ws_summary.cell(row=2, column=col)
            cell.font = header_font
            cell.fill = summary_header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Style summary data
        currency_rows = [3, 4, 5]  # Total Income, Expenses, Net Income rows
        category_header_row = 9  # "EXPENSE BREAKDOWN BY CATEGORY" row
        category_start_row = 10  # First category row

        for row in range(3, ws_summary.max_row + 1):
            # Metric label (column A)
            cell_a = ws_summary.cell(row=row, column=1)
            cell_b = ws_summary.cell(row=row, column=2)

            # Check if this is the category header row
            if row == category_header_row:
                cell_a.font = Font(name='Calibri', size=12, bold=True, color='1E293B')
                cell_a.alignment = cell_alignment
                cell_a.border = thin_border
                cell_a.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                cell_b.border = thin_border
                cell_b.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                continue

            # Check if this is a blank row
            if not cell_a.value or cell_a.value == '':
                cell_a.border = thin_border
                cell_b.border = thin_border
                continue

            # Check if this is a category item (indented with "  ")
            is_category_item = isinstance(cell_a.value, str) and cell_a.value.startswith('  ')

            if is_category_item:
                cell_a.font = Font(name='Calibri', size=10, color='334155')
            else:
                cell_a.font = metric_label_font

            cell_a.alignment = cell_alignment
            cell_a.border = thin_border
            cell_a.fill = alt_row_fill if row % 2 == 0 else white_fill

            # Value (column B)
            cell_b.font = metric_value_font
            cell_b.alignment = number_alignment
            cell_b.border = thin_border
            cell_b.fill = alt_row_fill if row % 2 == 0 else white_fill

            # Format currency values
            if row in currency_rows or is_category_item:
                cell_b.number_format = '$#,##0.00'
                # Highlight net income
                if row == 5:  # Net Income row
                    if isinstance(cell_b.value, (int, float)) and cell_b.value >= 0:
                        cell_b.fill = positive_fill
                    else:
                        cell_b.fill = negative_fill

        # Write income sheet
        if not income_df.empty:
            income_df.to_excel(writer, sheet_name='Income', index=False)
            ws_income = writer.sheets['Income']

            # Style headers
            for col in range(1, len(income_df.columns) + 1):
                cell = ws_income.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Style data rows
            for row in range(2, ws_income.max_row + 1):
                row_fill = alt_row_fill if row % 2 == 0 else white_fill
                for col in range(1, len(income_df.columns) + 1):
                    cell = ws_income.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.fill = row_fill

                    # Format based on column name
                    col_name = income_df.columns[col - 1].lower()
                    if 'amount' in col_name:
                        cell.number_format = '$#,##0.00'
                        cell.alignment = number_alignment
                    elif 'date' in col_name:
                        cell.number_format = 'yyyy-mm-dd'
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = cell_alignment

        # Write expenses sheet
        if not expenses_df.empty:
            expenses_df.to_excel(writer, sheet_name='Expenses', index=False)
            ws_expenses = writer.sheets['Expenses']

            # Style headers
            for col in range(1, len(expenses_df.columns) + 1):
                cell = ws_expenses.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Style data rows
            for row in range(2, ws_expenses.max_row + 1):
                row_fill = alt_row_fill if row % 2 == 0 else white_fill
                for col in range(1, len(expenses_df.columns) + 1):
                    cell = ws_expenses.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.fill = row_fill

                    # Format based on column name
                    col_name = expenses_df.columns[col - 1].lower()
                    if 'amount' in col_name:
                        cell.number_format = '$#,##0.00'
                        cell.alignment = number_alignment
                    elif 'date' in col_name:
                        cell.number_format = 'yyyy-mm-dd'
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = cell_alignment

        # Write expenses by category sheet
        if not expense_by_category_df.empty:
            expense_by_category_df.to_excel(writer, sheet_name='Expenses by Category', index=False)
            ws_category = writer.sheets['Expenses by Category']

            # Add header styling
            category_header_fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')

            # Style headers
            for col in range(1, len(expense_by_category_df.columns) + 1):
                cell = ws_category.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = category_header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Style data rows
            for row in range(2, ws_category.max_row + 1):
                row_fill = alt_row_fill if row % 2 == 0 else white_fill
                for col in range(1, len(expense_by_category_df.columns) + 1):
                    cell = ws_category.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.fill = row_fill

                    # Format based on column name
                    col_name = expense_by_category_df.columns[col - 1].lower()
                    if 'total' in col_name:
                        cell.number_format = '$#,##0.00'
                        cell.alignment = number_alignment
                    elif 'count' in col_name:
                        cell.number_format = '#,##0'
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = cell_alignment
                        cell.font = Font(name='Calibri', size=11, bold=True)

        # Write property summary sheet
        if not property_summary_df.empty:
            property_summary_df.to_excel(writer, sheet_name='Property Summary', index=False)
            ws_property = writer.sheets['Property Summary']

            # Style headers
            for col in range(1, len(property_summary_df.columns) + 1):
                cell = ws_property.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = property_header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Style data rows
            for row in range(2, ws_property.max_row + 1):
                row_fill = alt_row_fill if row % 2 == 0 else white_fill
                for col in range(1, len(property_summary_df.columns) + 1):
                    cell = ws_property.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.fill = row_fill

                    # Format based on column name
                    col_name = property_summary_df.columns[col - 1].lower()
                    if any(x in col_name for x in ['income', 'expense', 'net']):
                        cell.number_format = '$#,##0.00'
                        cell.alignment = number_alignment
                        # Highlight net column
                        if 'net' in col_name:
                            if isinstance(cell.value, (int, float)) and cell.value >= 0:
                                cell.fill = positive_fill
                            elif isinstance(cell.value, (int, float)):
                                cell.fill = negative_fill
                    elif 'transaction' in col_name:
                        cell.number_format = '#,##0'
                        cell.alignment = center_alignment
                    elif 'categories' in col_name or 'category' in col_name:
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        cell.font = Font(name='Calibri', size=9, color='334155')
                    elif 'property' in col_name:
                        cell.alignment = cell_alignment
                        cell.font = Font(name='Calibri', size=11, bold=True)
                    else:
                        cell.alignment = cell_alignment

        # Write property expense breakdown sheet
        if not property_expense_breakdown_df.empty:
            property_expense_breakdown_df.to_excel(writer, sheet_name='Property Expense Breakdown', index=False)
            ws_prop_expense = writer.sheets['Property Expense Breakdown']

            # Define a special fill for property expense breakdown
            expense_breakdown_header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')

            # Style headers
            for col in range(1, len(property_expense_breakdown_df.columns) + 1):
                cell = ws_prop_expense.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = expense_breakdown_header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Style data rows with property grouping
            current_property = None
            property_fill_toggle = True
            property_fill_1 = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            property_fill_2 = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')

            for row_idx in range(2, ws_prop_expense.max_row + 1):
                # Check if property changed (for visual grouping)
                property_cell = ws_prop_expense.cell(row=row_idx, column=1)
                if property_cell.value != current_property:
                    current_property = property_cell.value
                    property_fill_toggle = not property_fill_toggle

                # Determine fill color based on property grouping
                if property_fill_toggle:
                    base_fill = property_fill_1
                else:
                    base_fill = property_fill_2

                for col in range(1, len(property_expense_breakdown_df.columns) + 1):
                    cell = ws_prop_expense.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    cell.fill = base_fill

                    # Format based on column name
                    col_name = property_expense_breakdown_df.columns[col - 1].lower()
                    if 'amount' in col_name:
                        cell.number_format = '$#,##0.00'
                        cell.alignment = number_alignment
                        cell.font = Font(name='Calibri', size=10, bold=True, color='1E293B')
                    elif 'count' in col_name:
                        cell.number_format = '#,##0'
                        cell.alignment = center_alignment
                        cell.font = Font(name='Calibri', size=10, color='334155')
                    elif 'property' in col_name:
                        cell.alignment = cell_alignment
                        cell.font = Font(name='Calibri', size=11, bold=True, color='1E40AF')
                    elif 'category' in col_name:
                        cell.alignment = cell_alignment
                        cell.font = Font(name='Calibri', size=10, color='334155')
                    else:
                        cell.alignment = cell_alignment

        # Auto-adjust column widths for all sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    try:
                        # Skip merged cells
                        if hasattr(cell, 'column_letter'):
                            if column_letter is None:
                                column_letter = cell.column_letter
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                if column_letter:
                    adjusted_width = min(max(max_length + 3, 12), 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            # Freeze top row for data sheets
            if sheet_name != 'Summary':
                worksheet.freeze_panes = 'A2'

    buffer.seek(0)
    filename = f"lust_rentals_report_{resolved_year}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
